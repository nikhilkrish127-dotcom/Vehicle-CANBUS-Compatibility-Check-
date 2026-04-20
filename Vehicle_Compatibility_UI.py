
from pathlib import Path
import subprocess
import sys
import re
import os
from typing import List, Tuple

import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox

BASE = Path(__file__).resolve().parent
WORKBOOK = BASE / "Vehicle_Compatibility_Workbench_FINAL_v9_clean_report.xlsx"
SCRIPT = BASE / "vehicle_matcher_final_v11_clean_report.py"
FINAL_REPORT = BASE / "Final_Report_Clean.xlsx"

PREFERRED_DESC_HEADERS = ["raw_vehicle_description", "vehicle_description", "vehicle", "description", "asset_description"]
MAKE_HEADERS = ["make", "brand", "manufacturer", "oem"]
MODEL_HEADERS = ["model", "variant", "asset description", "asset_description", "type", "vehicle type", "vehicle_type", "body", "description"]
YEAR_HEADERS = ["year", "model year", "model_year", "year of manuf/ model year", "year_of_manuf_model_year", "manufacture year", "manufacture_year"]
EXTRA_HEADERS = ["generation", "series", "chassis", "trim", "fuel", "body", "notes"]

def norm_header(h):
    return re.sub(r"\s+", " ", str(h).strip().lower()).replace("/", " ").replace("-", " ").strip()

def clean_piece(v):
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    return s

def choose_best_sheet(path: Path):
    if path.suffix.lower() == ".csv":
        return "CSV", pd.read_csv(path)
    xls = pd.ExcelFile(path)
    best_name = None
    best_df = None
    best_score = -1
    for s in xls.sheet_names:
        try:
            df = pd.read_excel(path, sheet_name=s)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        non_empty_rows = int(df.dropna(how="all").shape[0])
        non_empty_cols = int(df.dropna(axis=1, how="all").shape[1])
        score = non_empty_rows * max(1, non_empty_cols)
        if score > best_score:
            best_score = score
            best_name = s
            best_df = df
    if best_df is None:
        raise ValueError("Could not find a usable sheet in the selected file.")
    return best_name, best_df


def detect_vehicle_strings(df: pd.DataFrame) -> Tuple[List[str], str]:
    if df is None or df.empty:
        return [], "Empty file"

    working = df.copy()
    working.columns = [norm_header(c) for c in working.columns]
    working = working.dropna(how="all")
    headers = list(working.columns)

    make_cols = [c for c in headers if any(k == norm_header(c) or k in norm_header(c) for k in MAKE_HEADERS)]
    model_cols = [c for c in headers if any(k == norm_header(c) or k in norm_header(c) for k in MODEL_HEADERS)]
    year_cols = [c for c in headers if any(k == norm_header(c) or k in norm_header(c) for k in YEAR_HEADERS)]
    extra_cols = [c for c in headers if any(k in norm_header(c) for k in EXTRA_HEADERS)]

    # 1) if structured columns exist, prefer composing from them
    if make_cols or model_cols or year_cols:
        vehicles = []
        for _, row in working.iterrows():
            parts = []
            seen = set()

            def add_piece(piece):
                piece = clean_piece(piece)
                if not piece:
                    return
                pnorm = piece.lower()
                if pnorm in seen:
                    return
                seen.add(pnorm)
                parts.append(piece)

            for c in make_cols[:2]:
                add_piece(row.get(c, ""))
            for c in model_cols[:3]:
                add_piece(row.get(c, ""))
            for c in extra_cols[:2]:
                val = clean_piece(row.get(c, ""))
                if val and val.lower() not in seen and len(val) <= 20:
                    add_piece(val)
            year_added = False
            for c in year_cols[:2]:
                val = clean_piece(row.get(c, ""))
                m = re.search(r"(19|20)\d{2}", val)
                if m:
                    add_piece(m.group(0))
                    year_added = True
                    break
            if not year_added:
                joined = " ".join(parts)
                m = re.search(r"(19|20)\d{2}", joined)
                if m and m.group(0).lower() not in seen:
                    add_piece(m.group(0))
            vehicle = re.sub(r"\s+", " ", " ".join(parts)).strip()
            if vehicle:
                vehicles.append(vehicle)
        if vehicles:
            cols_used = ", ".join(make_cols[:2] + model_cols[:3] + year_cols[:2])
            return vehicles, f"Composed from columns: {cols_used}"

    # 2) prefer a rich single description column
    rich_candidates = []
    for col in headers:
        col_norm = norm_header(col)
        vals = working[col].dropna().astype(str).str.strip()
        avg_len = float(vals.str.len().mean()) if not vals.empty else 0
        unique_ratio = float(vals.nunique() / max(1, len(vals))) if not vals.empty else 0
        year_ratio = float(vals.str.contains(r"(19|20)\d{2}", regex=True).mean()) if not vals.empty else 0
        score = 0
        if any(h in col_norm for h in PREFERRED_DESC_HEADERS):
            score += 8
        if "description" in col_norm:
            score += 4
        if "vehicle" in col_norm:
            score += 3
        if avg_len >= 12:
            score += 2
        if unique_ratio >= 0.5:
            score += 1
        if year_ratio >= 0.2:
            score += 2
        if score > 0:
            rich_candidates.append((score, col, avg_len))
    rich_candidates.sort(reverse=True)
    if rich_candidates:
        col = rich_candidates[0][1]
        vals = [clean_piece(v) for v in working[col].tolist()]
        vals = [v for v in vals if v]
        if vals:
            return vals, f"Used single description column: {col}"

    # 3) fallback: first non-empty text column
    for col in headers:
        vals = [clean_piece(v) for v in working[col].tolist()]
        vals = [v for v in vals if v]
        if vals:
            return vals, f"Fallback to first usable column: {col}"
    return [], "No usable vehicle column found"


class VehicleMatchUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Vehicle Compatibility Matcher")
        self.geometry("1380x820")
        self.minsize(1220, 720)
        self.import_meta = ""

        main = ttk.Frame(self, padding=12)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Enter one vehicle per line. Each row should contain: Brand Model Year of Make").grid(row=0, column=0, sticky="w")
        self.input_box = tk.Text(main, height=8, width=130, wrap="word")
        self.input_box.grid(row=1, column=0, columnspan=8, sticky="nsew", pady=(4,10))
        self.input_box.insert("1.0", "Toyota Yaris 2022\nMercedes-Benz Actros 3348 MP2 2005")

        btns = ttk.Frame(main)
        btns.grid(row=2, column=0, columnspan=8, sticky="w", pady=(0,10))
        ttk.Button(btns, text="Run Match", command=self.run_match).pack(side="left")
        ttk.Button(btns, text="Open Workbook", command=self.open_workbook).pack(side="left", padx=(8,0))
        ttk.Button(btns, text="Open Final Report", command=self.open_final_report).pack(side="left", padx=(8,0))
        ttk.Button(btns, text="Load Results", command=self.load_results).pack(side="left", padx=(8,0))
        ttk.Button(btns, text="Clear Input", command=lambda: self.input_box.delete("1.0", "end")).pack(side="left", padx=(8,0))

        self.info_var = tk.StringVar(value="Paste or type vehicle rows manually in the format: Brand Model Year of Make")
        ttk.Label(main, textvariable=self.info_var).grid(row=3, column=0, columnspan=8, sticky="w", pady=(0,8))

        self.notebook = ttk.Notebook(main)
        self.notebook.grid(row=4, column=0, columnspan=8, sticky="nsew")

        summary_frame = ttk.Frame(self.notebook, padding=6)
        self.notebook.add(summary_frame, text="Best Results")
        ttk.Label(summary_frame, text="Best result per vehicle").grid(row=0, column=0, sticky="w")
        self.summary_tree = ttk.Treeview(
            summary_frame,
            columns=("Input", "Status", "Confidence", "Best Source", "Matched Brand", "Matched Model", "Year Check", "Fuel"),
            show="headings",
            height=12
        )
        for col, width in [
            ("Input", 260), ("Status", 120), ("Confidence", 90), ("Best Source", 110),
            ("Matched Brand", 120), ("Matched Model", 320), ("Year Check", 120), ("Fuel", 90)
        ]:
            self.summary_tree.heading(col, text=col)
            self.summary_tree.column(col, width=width, anchor="w")
        self.summary_tree.grid(row=1, column=0, sticky="nsew")
        summary_scroll = ttk.Scrollbar(summary_frame, orient="vertical", command=self.summary_tree.yview)
        summary_scroll.grid(row=1, column=1, sticky="ns")
        self.summary_tree.configure(yscrollcommand=summary_scroll.set)
        summary_frame.grid_rowconfigure(1, weight=1)
        summary_frame.grid_columnconfigure(0, weight=1)

        source_frame = ttk.Frame(self.notebook, padding=6)
        self.notebook.add(source_frame, text="Results by Adaptor")
        ttk.Label(source_frame, text="Best result per adaptor for each vehicle").grid(row=0, column=0, sticky="w")
        self.source_tree = ttk.Treeview(
            source_frame,
            columns=("Input", "Source", "Status", "Confidence", "Matched Brand", "Matched Model", "Year Range", "Brand Check", "Model Check", "Year Check", "Fuel", "Param Count"),
            show="headings",
            height=16
        )
        for col, width in [
            ("Input", 240), ("Source", 120), ("Status", 120), ("Confidence", 90),
            ("Matched Brand", 120), ("Matched Model", 280), ("Year Range", 110),
            ("Brand Check", 90), ("Model Check", 90), ("Year Check", 110), ("Fuel", 80), ("Param Count", 90)
        ]:
            self.source_tree.heading(col, text=col)
            self.source_tree.column(col, width=width, anchor="w")
        self.source_tree.grid(row=1, column=0, sticky="nsew")
        source_scroll = ttk.Scrollbar(source_frame, orient="vertical", command=self.source_tree.yview)
        source_scroll.grid(row=1, column=1, sticky="ns")
        self.source_tree.configure(yscrollcommand=source_scroll.set)
        source_frame.grid_rowconfigure(1, weight=1)
        source_frame.grid_columnconfigure(0, weight=1)

        detail_frame = ttk.LabelFrame(main, text="Selected row details", padding=8)
        detail_frame.grid(row=5, column=0, columnspan=8, sticky="nsew", pady=(10,0))
        self.detail_text = tk.Text(detail_frame, height=9, wrap="word")
        self.detail_text.grid(row=0, column=0, sticky="nsew")
        detail_scroll = ttk.Scrollbar(detail_frame, orient="vertical", command=self.detail_text.yview)
        detail_scroll.grid(row=0, column=1, sticky="ns")
        self.detail_text.configure(yscrollcommand=detail_scroll.set)
        detail_frame.grid_rowconfigure(0, weight=1)
        detail_frame.grid_columnconfigure(0, weight=1)
        self.detail_text.insert("1.0", "Select a row to see the reason and parameter preview.")
        self.detail_text.config(state="disabled")

        self.summary_tree.bind("<<TreeviewSelect>>", self.show_summary_detail)
        self.source_tree.bind("<<TreeviewSelect>>", self.show_source_detail)

        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(main, textvariable=self.status_var).grid(row=6, column=0, columnspan=8, sticky="w", pady=(10,0))

        main.grid_rowconfigure(4, weight=1)
        main.grid_columnconfigure(0, weight=1)

        self.summary_rows = {}
        self.source_rows = {}

    def set_detail(self, msg):
        self.detail_text.config(state="normal")
        self.detail_text.delete("1.0", "end")
        self.detail_text.insert("1.0", msg)
        self.detail_text.config(state="disabled")

    def _open_path(self, path):
        try:
            path = Path(path)
            if not path.exists():
                messagebox.showerror("Open file", f"File not found:\n{path}")
                return
            if sys.platform == "darwin":
                subprocess.run(["open", str(path)], check=False)
            elif os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except Exception as e:
            messagebox.showerror("Open file", str(e))

    def open_workbook(self):
        self._open_path(WORKBOOK)

    def open_final_report(self):
        self._open_path(FINAL_REPORT)

    def _extract_vehicles_from_text(self):
        raw = self.input_box.get("1.0", "end").strip()
        raw = raw.replace("\u2028", "\n").replace("\u2029", "\n").replace("\r", "\n")
        lines = [re.sub(r'\s+', ' ', line).strip() for line in raw.split("\n") if line.strip()]
        vehicles = []
        buffer = ""
        for line in lines:
            if re.fullmatch(r'\d{4}', line):
                if buffer:
                    vehicles.append(f"{buffer} {line}".strip())
                    buffer = ""
                elif vehicles:
                    vehicles[-1] = f"{vehicles[-1]} {line}".strip()
                else:
                    vehicles.append(line)
            elif re.search(r'\b\d{4}\b$', line):
                if buffer:
                    vehicles.append(f"{buffer} {line}".strip())
                    buffer = ""
                else:
                    vehicles.append(line)
            else:
                if buffer:
                    buffer = f"{buffer} {line}".strip()
                else:
                    buffer = line
        if buffer:
            vehicles.append(buffer)
        return vehicles

    def _write_inputs_to_workbook(self, vehicles):
        wb = load_workbook(WORKBOOK)
        ws = wb["Vehicle_Input"]
        max_clear = max(ws.max_row, len(vehicles) + 10, 1000)
        for r in range(2, max_clear + 1):
            ws[f"A{r}"] = None
            ws[f"B{r}"] = None
            ws[f"C{r}"] = None
            ws[f"D{r}"] = None
            ws[f"E{r}"] = None
        for idx, vehicle in enumerate(vehicles, start=2):
            ws[f"A{idx}"] = idx - 1
            ws[f"B{idx}"] = vehicle
            ws[f"C{idx}"] = self.import_meta
        wb.save(WORKBOOK)

    def run_match(self):
        vehicles = self._extract_vehicles_from_text()
        if not vehicles:
            messagebox.showwarning("Missing input", "Enter at least one vehicle description. Use one vehicle per row in the format Brand Model Year of Make.")
            return
        if not WORKBOOK.exists():
            messagebox.showerror("Missing workbook", f"Workbook not found:\n{WORKBOOK}")
            return
        if not SCRIPT.exists():
            messagebox.showerror("Missing script", f"Matcher script not found:\n{SCRIPT}")
            return
        try:
            self._write_inputs_to_workbook(vehicles)
            self.status_var.set("Running smart matcher...")
            self.update_idletasks()
            result = subprocess.run(
                [sys.executable, str(SCRIPT)],
                cwd=str(BASE),
                check=True,
                capture_output=True,
                text=True
            )
            msg = (result.stdout or "Completed.").strip().splitlines()
            self.status_var.set(msg[-1] if msg else "Completed.")
            self.load_results()
        except subprocess.CalledProcessError as e:
            self.status_var.set("Matcher failed.")
            messagebox.showerror("Matcher failed", (e.stderr or e.stdout or str(e)))
        except Exception as e:
            self.status_var.set("Error.")
            messagebox.showerror("Error", str(e))

    def load_results(self):
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for item in self.source_tree.get_children():
            self.source_tree.delete(item)
        self.summary_rows = {}
        self.source_rows = {}

        try:
            df = pd.read_excel(WORKBOOK, sheet_name="Match_Results")
            df = df[df["Raw_Vehicle_Description"].notna()]
            for _, r in df.iterrows():
                iid = self.summary_tree.insert("", "end", values=(
                    str(r.get("Raw_Vehicle_Description", "")),
                    str(r.get("Best_Status", "")),
                    str(r.get("Confidence", "")),
                    str(r.get("Best_Source", "")),
                    str(r.get("Matched_Brand", "")),
                    str(r.get("Matched_Model", "")),
                    str(r.get("Year_Check", "")),
                    str(r.get("Fuel_Data_Available", "")),
                ))
                self.summary_rows[iid] = {
                    "reason": str(r.get("Overall_Reason", "")),
                    "action": str(r.get("Suggested_Action", "")),
                    "preview": str(r.get("Parameter_Preview", "")),
                    "count": str(r.get("Supported_Parameter_Count", "")),
                }

            ps = pd.read_excel(WORKBOOK, sheet_name="Per_Source_Results")
            ps = ps[ps["Raw_Vehicle_Description"].notna()]
            for _, r in ps.iterrows():
                iid = self.source_tree.insert("", "end", values=(
                    str(r.get("Raw_Vehicle_Description", "")),
                    str(r.get("Source", "")),
                    str(r.get("Status", "")),
                    str(r.get("Confidence", "")),
                    str(r.get("Matched_Brand", "")),
                    str(r.get("Matched_Model", "")),
                    str(r.get("Supported_Year_Range", "")),
                    str(r.get("Brand_Check", "")),
                    str(r.get("Model_Check", "")),
                    str(r.get("Year_Check", "")),
                    str(r.get("Fuel_Data_Available", "")),
                    str(r.get("Supported_Parameter_Count", "")),
                ))
                self.source_rows[iid] = {
                    "reason": str(r.get("Reason", "")),
                    "reason_detail": str(r.get("Reason_Detail", "")),
                    "preview": str(r.get("Parameter_Preview", "")),
                    "full_params": str(r.get("Full_Supported_Parameters", "")),
                    "fuel": str(r.get("Fuel_Data_Available", "")),
                    "count": str(r.get("Supported_Parameter_Count", "")),
                }

            self.status_var.set(f"Loaded {len(df)} best result row(s) and {len(ps)} adaptor-result row(s).")
            self.set_detail("Select a row to see the full reason and parameter preview.")
        except Exception as e:
            messagebox.showerror("Load results", str(e))

    def show_summary_detail(self, event=None):
        selected = self.summary_tree.selection()
        if not selected:
            return
        info = self.summary_rows.get(selected[0], {})
        msg = (
            f"Reason summary: {info.get('reason','')}\n\n"
            f"Suggested action: {info.get('action','')}\n\n"
            f"Supported parameter count: {info.get('count','')}\n\n"
            f"Parameter preview:\n- {str(info.get('preview','')).replace('; ', chr(10) + '- ')}"
        )
        self.set_detail(msg)

    def show_source_detail(self, event=None):
        selected = self.source_tree.selection()
        if not selected:
            return
        info = self.source_rows.get(selected[0], {})
        params_block = info.get("full_params","") or info.get("preview","")
        msg = (
            f"Reason summary: {info.get('reason','')}\n\n"
            f"Detailed reason: {info.get('reason_detail','')}\n\n"
            f"Fuel data available: {info.get('fuel','')}\n"
            f"Supported parameter count: {info.get('count','')}\n\n"
            f"Supported parameters:\n{params_block.replace('; ', chr(10) + '- ')}"
        )
        if params_block:
            msg = msg.replace("Supported parameters:\n", "Supported parameters:\n- ", 1)
        self.set_detail(msg)

if __name__ == "__main__":
    VehicleMatchUI().mainloop()
